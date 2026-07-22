"""New Hampshire agency-level budget shifts.

Source: NH Office of Legislative Budget Assistant, HB 1 final (chaptered)
line-item Excel files for the FY2024-25 and FY2026-27 biennia.
Aggregates TYPE=E (expenditure) line items by DEPARTMENT NAME.

Download into scripts/nasbo/raw/ first:
  nh_hb1_final_2627.xlsx  https://gc.nh.gov/lba/budget/operating_budgets/2026-2027/House_Finance/Final%20HB%201%20-%20Excel%2007-22-25.xlsx
  nh_hb1_final_2425.xlsx  https://gc.nh.gov/lba/budget/operating_budgets/2024-2025/House_Finance/HB%201%20Final%20(FY%202024%20FY%202025).xlsx
"""
import json
import os
from collections import defaultdict

import openpyxl

DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "raw")
REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def dept_totals(fname):
    wb = openpyxl.load_workbook(os.path.join(DIR, fname), read_only=True)
    ws = wb[wb.sheetnames[0]]
    out = defaultdict(lambda: [0.0, 0.0])
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[10] != "E":  # E = expenditure lines; F mirrors them as funding sources
            continue
        dept = (row[3] or "").strip()
        if not dept:
            continue
        out[dept][0] += float(row[13] or 0)
        out[dept][1] += float(row[14] or 0)
    wb.close()
    return out


def title(s):
    small = {"OF", "THE", "AND", "FOR", "TO", "ON"}
    fixes = {"Dept": "Dept.", "Svs": "Services", "Hhs": "HHS", "It": "IT",
             "Nh": "NH", "Bea": "BEA", "Osi": "OSI"}
    words = []
    for i, w in enumerate(s.title().split()):
        u = w.upper()
        if i > 0 and u in small:
            words.append(w.lower())
        else:
            words.append(fixes.get(w, w))
    return " ".join(words)


cur = dept_totals("nh_hb1_final_2627.xlsx")
pri = dept_totals("nh_hb1_final_2425.xlsx")

agencies = []
for dept in sorted(set(cur) | set(pri)):
    c = cur.get(dept, [0, 0])
    p = pri.get(dept, [0, 0])
    agencies.append({
        "name": title(dept),
        "prior": round(p[0] + p[1]),
        "current": round(c[0] + c[1]),
        "years": {
            "fy2024": round(p[0]), "fy2025": round(p[1]),
            "fy2026": round(c[0]), "fy2027": round(c[1]),
        },
    })

out = {
    "state": "New Hampshire",
    "updated": "2026-07-17",
    "cycle": "biennial",
    "basis": "All funds, legislative appropriations (HB 1 operating budget expenditure line items, TYPE=E)",
    "priorLabel": "FY2024-25 biennium (enacted 2023)",
    "currentLabel": "FY2026-27 biennium (enacted Jun 2025, chaptered final)",
    "unit": "dollars",
    "sources": [
        {"label": "NH LBA, HB 1 Final FY2026-27 (Excel, chaptered 7/22/25)",
         "url": "https://gc.nh.gov/lba/budget/fy2026_2027_budget.aspx"},
        {"label": "NH LBA, HB 1 Final FY2024-25 (Excel)",
         "url": "https://gc.nh.gov/lba/budget/fy2024_2025_budget.aspx"},
    ],
    "notes": [
        "Amounts are biennium totals of legislative appropriations, all fund sources. Funding-source (TYPE=F) rows mirror expenditure totals, providing an internal consistency check.",
        "Departments new or reorganized between biennia show as large swings; NH reorganized several agencies in the 2026-27 budget.",
    ],
    "agencies": agencies,
}

dest = os.path.join(REPO, "data", "budget-agencies", "new-hampshire.json")
os.makedirs(os.path.dirname(dest), exist_ok=True)
with open(dest, "w", encoding="utf-8") as f:
    json.dump(out, f, indent=1, ensure_ascii=False)

tp = sum(a["prior"] for a in agencies)
tc = sum(a["current"] for a in agencies)
print("Wrote", dest)
print("agencies:", len(agencies))
print("prior total: $%.2fB  current total: $%.2fB  (%+.1f%%)" % (tp / 1e9, tc / 1e9, (tc / tp - 1) * 100))
for a in sorted(agencies, key=lambda x: -(x["current"] - x["prior"]))[:5]:
    print("  +", a["name"], "%+.0fM" % ((a["current"] - a["prior"]) / 1e6))
for a in sorted(agencies, key=lambda x: (x["current"] - x["prior"]))[:5]:
    print("  -", a["name"], "%+.0fM" % ((a["current"] - a["prior"]) / 1e6))
