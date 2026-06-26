# Parse the BDM + Top Revenue tabs of the SLG Focused Accounts workbook into a
# clean, per-state, public-safe account list. Internal fields (revenue, acct id,
# BDM rep) are used only for selection/prioritization, never written to the manifest.
import openpyxl, os, json, re

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
wb = openpyxl.load_workbook(os.path.join(ROOT, "SLG Focused Accounts.xlsx"), data_only=True, read_only=True)

STATE_ABBR = {
 'AL':'alabama','AK':'alaska','AZ':'arizona','AR':'arkansas','CA':'california','CO':'colorado','CT':'connecticut',
 'DE':'delaware','FL':'florida','GA':'georgia','HI':'hawaii','ID':'idaho','IL':'illinois','IN':'indiana','IA':'iowa',
 'KS':'kansas','KY':'kentucky','LA':'louisiana','ME':'maine','MD':'maryland','MA':'massachusetts','MI':'michigan',
 'MN':'minnesota','MS':'mississippi','MO':'missouri','MT':'montana','NE':'nebraska','NV':'nevada','NH':'new-hampshire',
 'NJ':'new-jersey','NM':'new-mexico','NY':'new-york','NC':'north-carolina','ND':'north-dakota','OH':'ohio','OK':'oklahoma',
 'OR':'oregon','PA':'pennsylvania','RI':'rhode-island','SC':'south-carolina','SD':'south-dakota','TN':'tennessee',
 'TX':'texas','UT':'utah','VT':'vermont','VA':'virginia','WA':'washington','WV':'west-virginia','WI':'wisconsin','WY':'wyoming'}
NAME_TO_ABBR = {
 'ALABAMA':'AL','ALASKA':'AK','ARIZONA':'AZ','ARKANSAS':'AR','CALIFORNIA':'CA','COLORADO':'CO','CONNECTICUT':'CT',
 'DELAWARE':'DE','FLORIDA':'FL','GEORGIA':'GA','HAWAII':'HI','IDAHO':'ID','ILLINOIS':'IL','INDIANA':'IN','IOWA':'IA',
 'KANSAS':'KS','KENTUCKY':'KY','LOUISIANA':'LA','MAINE':'ME','MARYLAND':'MD','MASSACHUSETTS':'MA','MICHIGAN':'MI',
 'MINNESOTA':'MN','MISSISSIPPI':'MS','MISSOURI':'MO','MONTANA':'MT','NEBRASKA':'NE','NEVADA':'NV','NEW HAMPSHIRE':'NH',
 'NEW JERSEY':'NJ','NEW MEXICO':'NM','NEW YORK':'NY','NORTH CAROLINA':'NC','NORTH DAKOTA':'ND','OHIO':'OH','OKLAHOMA':'OK',
 'OREGON':'OR','PENNSYLVANIA':'PA','RHODE ISLAND':'RI','SOUTH CAROLINA':'SC','SOUTH DAKOTA':'SD','TENNESSEE':'TN',
 'TEXAS':'TX','UTAH':'UT','VERMONT':'VT','VIRGINIA':'VA','WASHINGTON':'WA','WEST VIRGINIA':'WV','WISCONSIN':'WI','WYOMING':'WY'}

def norm_region(r):
    if not r: return ''
    s = re.sub(r'\s+', ' ', str(r)).strip().lower().replace('region','').strip(' 0123456789-')
    m = {'northeast':'Northeast','south east':'Southeast','southeast':'Southeast','midwest':'Midwest',
         'central':'Central','central west':'Central West','west':'West'}
    return m.get(s, str(r).strip())

def infer_state(name):
    u = ' ' + re.sub(r'[^A-Z ]', ' ', name.upper()) + ' '
    if 'DISTRICT OF COLUMBIA' in name.upper(): return 'DC'
    # explicit "STATE OF XX" or trailing ", XX"
    for full, ab in NAME_TO_ABBR.items():
        if (' ' + full + ' ') in u:
            return ab
    m = re.search(r'\b([A-Z]{2})\b\s*$', name.strip())
    if m and m.group(1) in STATE_ABBR: return m.group(1)
    m = re.search(r'STATE OF ([A-Z]{2})\b', name.upper())
    if m and m.group(1) in STATE_ABBR: return m.group(1)
    return None

def classify(name):
    u = name.upper()
    if any(k in u for k in ['TRANSIT','AIRPORT','PORT AUTHORITY','METRO ','METROPOLITAN','TURNPIKE','TOLL','WATER AUTHORITY','HOUSING AUTHORITY']):
        return 'Special District'
    if 'COURT' in u: return 'Court'
    if 'COUNTY' in u: return 'County'
    if 'SCHOOL DISTRICT' in u or ' ISD' in u or ' USD' in u or 'PUBLIC SCHOOLS' in u: return 'K-12'
    if 'UNIVERSITY' in u or 'COLLEGE' in u or 'REGENTS' in u: return 'Higher Ed'
    if u.startswith('NYC') or 'CITY OF' in u: return 'City'
    # "CityName, FullStateName" with no agency keyword -> City
    if ',' in name:
        tail = name.split(',')[-1].strip().upper()
        if tail in NAME_TO_ABBR and not any(k in u for k in ['STATE OF','DEPARTMENT','DEPT','OFFICE','COMMISSION','AGENCY','DIVISION','AUTHORITY','SERVICES','SYSTEM']):
            return 'City'
    if any(k in u for k in ['STATE OF','COMMONWEALTH','DEPARTMENT','DEPT','OFFICE OF','COMMISSION','CABINET','AGENCY','DIVISION','AUTHORITY','SERVICES','DTMB','MNIT','OIT ','DTS','DIR','DOIT','BUREAU','TREASURER','ATTORNEY GENERAL','POLICE','NYPD']):
        return 'State Agency'
    return 'Other'

ACR = set(STATE_ABBR.keys()) | set(
 'NYC NYPD HHS HHSC HHSA CHHS DIR OIT OITS DTS DTMB MNIT DOIT IOT ITSD ITS IT DC US AG DOT DGS BIT '
 'COT OMES DAS OA ETS VITA EOTSS COVA DPS DMV CIO CISO MTA RHT NASBO DTI ADOA SITSD EIS WATECH OFM'.split())
SMALL = set('of and the for a an to as at by in on or & de la'.split())
def titlecase(name):
    name = re.sub(r'\s+', ' ', name.replace('|', ' — ')).strip()
    words = name.split(' ')
    out = []
    for i, w in enumerate(words):
        m = re.match(r'^([(\[\"\'#]*)(.*?)([)\]\"\',.;:]*)$', w)
        pre, core, post = m.group(1), m.group(2), m.group(3)
        low = core.lower()
        if not core:
            res = w
        elif core.upper() in ACR:
            res = pre + core.upper() + post
        elif low in SMALL and 0 < i < len(words) - 1:
            res = pre + low + post
        else:
            res = pre + (core[:1].upper() + core[1:].lower() if core.isalpha() else core) + post
        out.append(res)
    return ' '.join(out)

records = {}  # acctId -> dict
def add(aid, name, region, source, state=None, rev=None):
    aid = str(aid).strip()
    if not re.match(r'^\d+$', aid): return
    r = records.setdefault(aid, {'names': [], 'region': '', 'state': None, 'sources': set(), 'rev': None})
    if name: r['names'].append(name.strip())
    if region and not r['region']: r['region'] = norm_region(region)
    if state and not r['state']: r['state'] = state
    if rev is not None: r['rev'] = rev
    r['sources'].add(source)

# ---- Top Revenue tab (authoritative state) ----
ws = wb['SLG Top Revenue Accounts']
rows = [list(r) for r in ws.iter_rows(values_only=True)]
def cell(rr, cc):
    return rows[rr][cc] if 0 <= rr < len(rows) and cc < len(rows[rr]) else None
for r in range(len(rows)):
    for c in range(len(rows[r])):
        if str(cell(r, c)).strip() == 'Rank':
            # region = nearest 'Region' text above in these columns
            region = ''
            for rr in range(r-1, -1, -1):
                for cc in range(c, min(c+7, len(rows[rr]))):
                    v = cell(rr, cc)
                    if v and 'Region' in str(v):
                        region = str(v); break
                if region: break
            rr = r + 1
            while rr < len(rows):
                rank = cell(rr, c); aid = cell(rr, c+1); nm = cell(rr, c+2); st = cell(rr, c+3); rev = cell(rr, c+4)
                if not aid or not str(aid).strip(): break
                if str(aid).strip() in ('Rank', 'Account ID'): break
                add(aid, nm, region, 'TopRev', state=(str(st).strip().upper() if st else None), rev=rev)
                rr += 1

# ---- BDM tab ----
ws = wb['SLG BDM Focused Accounts']
rows = [list(r) for r in ws.iter_rows(values_only=True)]
for r in range(len(rows)):
    for c in range(len(rows[r])):
        if str(cell(r, c)).strip() == 'ACCT ID':
            region = cell(r, c+2)
            rr = r + 1
            while rr < len(rows):
                aid = cell(rr, c); nm = cell(rr, c+1); sub = cell(rr, c+2)
                if not aid or not re.match(r'^\d+$', str(aid).strip()): break
                full = str(nm).strip() if nm else ''
                if sub and str(sub).strip() and str(sub).strip().lower() not in ('northeast','southeast','south east','midwest','central','central west','west'):
                    full = full + ' | ' + str(sub).strip()
                add(aid, full, region, 'BDM')
                rr += 1

# ---- consolidate ----
by_state = {}
unmapped = []
type_counts = {}
for aid, r in records.items():
    name = max(r['names'], key=len) if r['names'] else ''
    if not name: continue
    abbr = r['state'] or infer_state(name)
    disp = titlecase(name)
    typ = classify(name)
    type_counts[typ] = type_counts.get(typ, 0) + 1
    rec = {'type': typ, 'name': disp, 'region': r['region'], 'sources': sorted(r['sources']), 'rev': r['rev'], 'raw': name}
    if abbr and abbr in STATE_ABBR:
        by_state.setdefault(STATE_ABBR[abbr], []).append(rec)
    else:
        unmapped.append({**rec, 'abbr': abbr})

def slugify(s, used):
    base = re.sub(r'[^a-z0-9]+', '-', s.lower()).strip('-')[:48] or 'account'
    sl = base; i = 2
    while sl in used: sl = f"{base}-{i}"; i += 1
    used.add(sl); return sl

# Overrides: keep slugs/availability stable for accounts that already have briefs.
OVERRIDES = {
    'texas': [
        ('INFORMATION RESOURCES', {'slug': 'dir', 'name': 'Texas DIR — Department of Information Resources'}),
        ('HEALTH & HUMAN SERVICES', {'slug': 'hhsc', 'name': 'Texas Health & Human Services Commission (HHSC)'}),
        ('DEPT OF TRANSPORTATION', {'slug': 'txdot', 'name': 'Texas Dept of Transportation (TxDOT)'}),
        ('DEPT OF PUBLIC SAFETY', {'slug': 'dps', 'name': 'Texas Dept of Public Safety (DPS)'}),
        ('FAMILY AND PROTECTIVE', {'slug': 'dfps', 'name': 'Texas Dept of Family & Protective Services (DFPS)'}),
        ('WORKFORCE COMMISSION', {'slug': 'twc', 'name': 'Texas Workforce Commission (TWC)'}),
        ('AUSTIN', {'slug': 'austin', 'name': 'Austin, Texas'}),
        ('HARRIS COUNTY', {'slug': 'harris-county', 'name': 'Harris County, Texas'}),
    ],
    'virginia': [('COVA EXECUTIVE BRANCH', {'slug': 'vita', 'name': 'Virginia — Secretary of Administration / VITA', 'type': 'State Agency'})],
    'hawaii': [('DEPT OF ED', {'slug': 'doe', 'name': 'Hawaii Dept of Education'})],
    'colorado': [('COLORADO EXECUTIVE BRANCH', {'slug': 'oit', 'name': "Colorado Governor's Office of Information Technology (OIT)"})],
    'arkansas': [('STATE OF ARKANSAS', {'slug': 'dis', 'name': 'Arkansas Division of Information Systems (DIS)'})],
    'oregon': [('ADMINISTRATIVE SERVICES, STATE OF OREGON', {'slug': 'das', 'name': 'Oregon Dept of Administrative Services (DAS/EIS)'})],
    'michigan': [('DTMB', {'slug': 'dtmb', 'name': 'Michigan DTMB — Dept of Technology, Management & Budget'})],
    'louisiana': [('DIVISION OF ADMINISTRATION', {'slug': 'division-of-administration', 'name': 'Louisiana Division of Administration (OTS)'})],
    'new-york': [
        ('INFORMATION TECHNOLOGY AND TELECOMMUNICATIONS', {'slug': 'nyc-oti', 'name': 'NYC Office of Technology & Innovation (OTI)'}),
        ('METRO TRANSIT', {'slug': 'mta', 'name': 'NY Metropolitan Transportation Authority (MTA)'}),
        ('NYC HEALTH', {'slug': 'nyc-health', 'name': 'NYC Dept of Health & Mental Hygiene'}),
        ('NYPD', {'slug': 'nypd', 'name': 'New York City Police Department (NYPD)', 'type': 'City'}),
        ('NEW YORK STATE COURT SYSTEM', {'slug': 'nys-courts', 'name': 'New York State Court System'}),
        ('NYS OFFICE OF INFORMATION TECHNOLOGY', {'slug': 'nys-its', 'name': 'NYS Office of Information Technology Services (ITS)'}),
        ('HEALTH DEPARTMENT, STATE OF NY', {'slug': 'nys-doh', 'name': 'New York State Dept of Health'}),
        ('ENVIRONMENTAL PROTECTION', {'slug': 'nyc-dep', 'name': 'NYC Dept of Environmental Protection (DEP)', 'type': 'City'}),
    ],
    'tennessee': [
        ('GENERAL SERVICES', {'slug': 'general-services', 'name': 'Tennessee Dept of General Services'}),
        ('LAW, SAFETY', {'slug': 'safety-transportation', 'name': 'Tennessee — Safety, Corrections & Transportation', 'type': 'State Agency'}),
    ],
    'pennsylvania': [
        ('PHILADELPHIA', {'slug': 'philadelphia', 'name': 'Philadelphia, Pennsylvania'}),
        ('COMMONWEALTH OF PA 3', {'slug': 'oa-oit', 'name': 'Commonwealth of Pennsylvania (Office of Administration / OA-OIT)'}),
    ],
    'california': [
        ('LOS ANGELES COUNTY', {'slug': 'los-angeles-county', 'name': 'Los Angeles County, California'}),
        ('LOS ANGELES, CALIFORNIA', {'slug': 'los-angeles', 'name': 'Los Angeles, California', 'type': 'City'}),
        ('CHHS', {'slug': 'chhs', 'name': 'California Health & Human Services Agency (CHHS)'}),
        ('GENERAL SERVICES DEPARTMENT', {'slug': 'dgs', 'name': 'California Dept of General Services (DGS)'}),
        ('LEGISLATIVE DATA CENTER', {'slug': 'legislative-data-center', 'name': 'California Legislative Data Center', 'type': 'State Agency'}),
    ],
    'kentucky': [
        ('FINANCE & ADMINISTRATION', {'slug': 'finance-administration-cabinet', 'name': 'Kentucky Finance & Administration Cabinet'}),
        ('LOUISVILLE', {'slug': 'louisville-metro', 'name': 'Louisville/Jefferson County Metro, Kentucky'}),
    ],
    'delaware': [('GOVERNOR', {'slug': 'dti', 'name': 'Delaware Dept of Technology & Information (DTI)'})],
    'alabama': [('DEPT OF FINANCE', {'slug': 'finance', 'name': 'Alabama Dept of Finance (OIT)'})],
    'utah': [('DEPARTMENT OF TECHNOLOGY SERVICES', {'slug': 'dts', 'name': 'Utah Dept of Technology Services (DTS)'})],
    'oklahoma': [('MANAGEMENT AND ENTERPRISE SERVICES', {'slug': 'omes', 'name': 'Oklahoma Office of Management & Enterprise Services (OMES)'})],
    'minnesota': [('MN IT SERVICES', {'slug': 'mnit', 'name': 'Minnesota IT Services (MNIT)'})],
    'ohio': [('ADMINISTRATIVE SERVICES', {'slug': 'das', 'name': 'Ohio Dept of Administrative Services'})],
    'georgia': [('TECHNOLOGY AUTHORITY', {'slug': 'gta', 'name': 'Georgia Technology Authority (GTA)'})],
    'kansas': [('INFORMATION TECHNOLOGY SERVICES', {'slug': 'oits', 'name': 'Kansas Office of Information Technology Services'})],
    'north-carolina': [('ADMINISTRATION DEPARTMENT', {'slug': 'doa', 'name': 'North Carolina Dept of Administration'})],
    'washington': [
        ('INFORMATION SERVICES', {'slug': 'watech', 'name': 'Washington Technology Solutions (WaTech)'}),
        ('SOCIAL AND HEALTH', {'slug': 'dshs', 'name': 'Washington Dept of Social & Health Services (DSHS)'}),
    ],
    'nevada': [
        ('ADMINISTRATION DEPARTMENT', {'slug': 'doa', 'name': 'Nevada Dept of Administration'}),
        ('CLARK COUNTY', {'slug': 'clark-county', 'name': 'Clark County, Nevada'}),
    ],
    'maryland': [
        ('MONTGOMERY COUNTY', {'slug': 'montgomery-county', 'name': 'Montgomery County, Maryland'}),
        ('DEPARTMENT OF INFORMATION TECHNOLOGY', {'slug': 'doit', 'name': 'Maryland Dept of Information Technology (DoIT)'}),
        ('ATTORNEY GENERAL', {'slug': 'attorney-general', 'name': "Maryland Attorney General's Office"}),
    ],
    'illinois': [
        ('TOLL HIGHWAY', {'slug': 'tollway', 'name': 'Illinois Tollway (Toll Highway Authority)'}),
        ('CHICAGO', {'slug': 'chicago', 'name': 'Chicago, Illinois'}),
        ('DOIT', {'slug': 'doit', 'name': 'Illinois DoIT (Dept of Innovation & Technology)'}),
    ],
    'arizona': [
        ('MARICOPA COUNTY', {'slug': 'maricopa-county', 'name': 'Maricopa County, Arizona'}),
        ('PIMA COUNTY', {'slug': 'pima-county', 'name': 'Pima County, Arizona'}),
        ('STATE OF ARIZONA', {'slug': 'adoa', 'name': 'Arizona Dept of Administration (ADOA-ASET)'}),
    ],
    'florida': [
        ('HEALTH DEPARTMENT', {'slug': 'doh', 'name': 'Florida Dept of Health'}),
        ('TRANSPORTATION DEPARTMENT, STATE OF FL', {'slug': 'fdot', 'name': 'Florida Dept of Transportation (FDOT)'}),
        ('CHILDREN AND FAMILIES', {'slug': 'dcf', 'name': 'Florida Dept of Children & Families (DCF)'}),
        ('HIGHWAY SAFETY', {'slug': 'flhsmv', 'name': 'Florida Dept of Highway Safety & Motor Vehicles (FLHSMV)'}),
        ('MIAMI-DADE', {'slug': 'miami-dade-county', 'name': 'Miami-Dade County, Florida'}),
        ('CORRECTIONS DEPARTMENT', {'slug': 'fdc', 'name': 'Florida Dept of Corrections (FDC)'}),
        ('JACKSONVILLE', {'slug': 'jacksonville', 'name': 'Jacksonville (Duval County), Florida', 'type': 'City'}),
    ],
    'new-mexico': [('STATE OF NEW MEXICO', {'slug': 'doit', 'name': 'New Mexico Dept of Information Technology (DoIT)'})],
    'indiana': [('IOT', {'slug': 'iot', 'name': 'Indiana Office of Technology (IOT)'})],
    'new-jersey': [
        ('COURT SYSTEM', {'slug': 'courts', 'name': 'New Jersey Judiciary (Courts)'}),
        ('TREASURER', {'slug': 'treasury', 'name': 'New Jersey Dept of the Treasury'}),
        ('OFFICE OF INFORMATION TECHNOLOGY (OIT)', {'slug': 'njoit', 'name': 'New Jersey Office of Information Technology (NJOIT)'}),
        ('ATTORNEY GENERAL', {'slug': 'law-public-safety', 'name': 'New Jersey Attorney General / Dept of Law & Public Safety'}),
    ],
    'missouri': [('ITSD', {'slug': 'itsd', 'name': 'Missouri ITSD (OA Information Technology Services Division)'})],
    'alaska': [('INFORMATION TECHNOLOGY', {'slug': 'oit', 'name': 'Alaska Office of Information Technology (OIT)'})],
    'iowa': [('STATE OF IA', {'slug': 'ocio', 'name': 'Iowa Office of the Chief Information Officer (OCIO)'})],
    'south-carolina': [('SOCIAL SERVICES', {'slug': 'dss', 'name': 'South Carolina Dept of Social Services (DSS)'})],
}
# Accounts intentionally NOT given their own page — duplicates of an existing
# report. Matched against the raw account name (upper). Excluded from the
# manifest and the Built count so they don't show as a perpetual gap.
SKIP = {
    # "Commonwealth of PA 2" is just an odd label for the enterprise State of PA;
    # already covered by the Pennsylvania state report (+ the oa-oit account page).
    'pennsylvania': ['COMMONWEALTH OF PA 2'],
}

manifest = {}
worklist = []
for slug in sorted(by_state):
    used = set()
    items = sorted(by_state[slug], key=lambda x: (-(x['rev'] or 0), x['name']))
    out = []
    for it in items:
        if any(s in it['raw'].upper() for s in SKIP.get(slug, [])):
            continue
        e = {'type': it['type'], 'name': it['name'], 'slug': None, 'region': it['region'], 'available': False}
        for matchstr, data in OVERRIDES.get(slug, []):
            if matchstr in it['raw'].upper():
                e['name'] = data.get('name', e['name'])
                e['type'] = data.get('type', e['type'])
                e['slug'] = data.get('slug')
                break
        if e['slug']:
            used.add(e['slug'])
        else:
            e['slug'] = slugify(e['name'], used)
        # "available" is driven by whether a brief data file exists.
        e['available'] = os.path.exists(os.path.join(ROOT, 'data', 'accounts', slug, e['slug'] + '.json'))
        out.append(e)
        worklist.append((slug, e['slug'], e['name'], e['type'], it['rev'] or 0, e['available']))
    manifest[slug] = out

with open(os.path.join(ROOT, 'data', 'focused-accounts.json'), 'w', encoding='utf-8') as f:
    json.dump(manifest, f, indent=2, ensure_ascii=False)

# ---- summary ----
total = sum(len(v) for v in by_state.values())
print(f"Unique accounts: {len(records)}  |  mapped to states: {total}  |  unmapped: {len(unmapped)}")
print(f"States covered: {len(by_state)}")
print("\nType distribution:")
for t, n in sorted(type_counts.items(), key=lambda x: -x[1]): print(f"  {t:16} {n}")
print("\nAccounts per state:")
for slug in sorted(by_state, key=lambda s: -len(by_state[s])):
    print(f"  {slug:16} {len(by_state[slug])}")
print("\nUnmapped (no state page — e.g., DC / federal / ambiguous):")
for u in unmapped: print(f"  [{u.get('abbr')}] {u['name']}  ({','.join(u['sources'])})")
built = sum(1 for w in worklist if w[5])
print(f"\nBuilt: {built} / {len(worklist)}")
print("\nNOT-YET-BUILT, ranked by 3-yr revenue (state | slug | type | name):")
for slug, ac, nm, ty, rev, av in sorted(worklist, key=lambda x: -x[4]):
    if not av:
        print(f"  ${rev:>5}M  {slug:16} {ac:34} {ty:14} {nm}")
wb.close()
